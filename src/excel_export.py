from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.export_common import EXCEL_COLUMN_WIDTHS, EXPORT_FIELD_LABELS, record_export_value
from src.models import DeviceRecord


def export_records(
    path: Path,
    records: Iterable[DeviceRecord],
    fields: list[str],
) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Thiết bị Apple"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    for col, key in enumerate(fields, start=1):
        cell = ws.cell(row=1, column=col, value=EXPORT_FIELD_LABELS[key])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row_idx = 2
    for rec in records:
        for col, key in enumerate(fields, start=1):
            ws.cell(row=row_idx, column=col, value=record_export_value(rec, key))
        row_idx += 1

    for i, key in enumerate(fields, start=1):
        ws.column_dimensions[get_column_letter(i)].width = EXCEL_COLUMN_WIDTHS.get(key, 14)

    ws.freeze_panes = "A2"
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path
